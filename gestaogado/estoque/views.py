# pyright: reportAttributeAccessIssue=false
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, Avg
from .models import Produto, Movimentacao_estoque, MovimentacaoLote
from fazendascdst.models import Fazenda, Lote
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import io
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

@login_required
def index(request):
    produtos = Produto.objects.filter(user=request.user, quantidade__gt=0).select_related('fazenda')
    
    # Cálculos para estatísticas
    total_produtos = produtos.aggregate(total=Sum('quantidade'))['total'] or 0
    
    # Calcular valor total do estoque corretamente
    valor_total_estoque = 0
    for produto in produtos:
        valor_total_estoque += produto.quantidade * produto.valor_unitario
    
    # Produtos com baixo estoque (menos de 10 unidades)
    baixo_estoque = produtos.filter(quantidade__lt=10).count()
    
    # Categorias únicas
    categorias = produtos.values('tipo').distinct().count()
    
    # Produtos por categoria
    produtos_por_categoria = produtos.values('tipo').annotate(
        total=Count('id'),
        quantidade_total=Sum('quantidade')
    )
    
    context = {
        'produtos': produtos,
        'total_produtos': total_produtos,
        'valor_total_estoque': valor_total_estoque,
        'baixo_estoque': baixo_estoque,
        'categorias': categorias,
        'produtos_por_categoria': produtos_por_categoria,
    }
    
    return render(request, 'estoque/index.html', context)

@login_required
def adicionar_produto(request):
    if request.method == 'POST':
        try:
            fazenda_id = request.POST.get('fazenda')
            nome = request.POST.get('nome')
            tipo = request.POST.get('tipo')
            quantidade = int(request.POST.get('quantidade', 0))
            valor_unitario = float(request.POST.get('valor_unitario', 0))
            
            fazenda = Fazenda.objects.get(id=fazenda_id, user=request.user)
            
            produto = Produto.objects.create(
                fazenda=fazenda,
                nome=nome,
                tipo=tipo,
                quantidade=quantidade,
                valor_unitario=valor_unitario,
                user=request.user
            )
            
            # Registrar movimentação de entrada
            if quantidade > 0:
                Movimentacao_estoque.objects.create(
                    produto=produto,
                    tipo='entrada',
                    quantidade=quantidade,
                    data=request.POST.get('data', None),
                    observacao='Entrada inicial do produto',
                    user=request.user
                )
            
            messages.success(request, f'Produto {nome} adicionado com sucesso!')
            return redirect('estoque_index')
            
        except Exception as e:
            messages.error(request, f'Erro ao adicionar produto: {str(e)}')
    
    fazendas = Fazenda.objects.filter(user=request.user)
    return render(request, 'estoque/adicionar_produto.html', {'fazendas': fazendas})

@login_required
def editar_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id, user=request.user)
    
    if request.method == 'POST':
        try:
            produto.nome = request.POST.get('nome')
            produto.tipo = request.POST.get('tipo')
            produto.valor_unitario = float(request.POST.get('valor_unitario', 0))
            produto.save()
            
            messages.success(request, f'Produto {produto.nome} atualizado com sucesso!')
            return redirect('estoque_index')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar produto: {str(e)}')
    
    fazendas = Fazenda.objects.filter(user=request.user)
    return render(request, 'estoque/editar_produto.html', {
        'produto': produto,
        'fazendas': fazendas
    })

@login_required
def movimentar_estoque(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id, user=request.user)
    
    if request.method == 'POST':
        try:
            tipo_movimentacao = request.POST.get('tipo')
            quantidade = int(request.POST.get('quantidade', 0))
            data = request.POST.get('data')
            observacao = request.POST.get('observacao', '')
            
            if tipo_movimentacao == 'saida' and quantidade > produto.quantidade:
                messages.error(request, 'Quantidade insuficiente em estoque!')
                return redirect('estoque_index')
            
            # Registrar movimentação
            Movimentacao_estoque.objects.create(
                produto=produto,
                tipo=tipo_movimentacao,
                quantidade=quantidade,
                data=data,
                observacao=observacao,
                user=request.user
            )
            
            # Atualizar quantidade do produto
            if tipo_movimentacao == 'entrada':
                produto.quantidade += quantidade
            else:
                produto.quantidade -= quantidade
            
            produto.save()
            
            messages.success(request, f'Movimentação registrada com sucesso!')
            return redirect('estoque_index')
            
        except Exception as e:
            messages.error(request, f'Erro ao registrar movimentação: {str(e)}')
    
    movimentacoes = Movimentacao_estoque.objects.filter(produto=produto, user=request.user).order_by('-data')
    return render(request, 'estoque/movimentar_estoque.html', {
        'produto': produto,
        'movimentacoes': movimentacoes
    })

@login_required
def deletar_produto(request, produto_id):
    if request.method == 'POST':
        try:
            produto = Produto.objects.get(id=produto_id, user=request.user)
            nome = produto.nome
            produto.delete()
            messages.success(request, f'Produto {nome} removido com sucesso!')
        except Produto.DoesNotExist:
            messages.error(request, 'Produto não encontrado!')
    
    return redirect('estoque_index')

@login_required
def movimentacao_lote(request):
    if request.method == 'POST':
        try:
            lote_id = request.POST.get('lote')
            tipo = request.POST.get('tipo')
            quantidade = int(request.POST.get('quantidade', 0))
            peso_medio = float(request.POST.get('peso_medio', 0))
            data = request.POST.get('data') or timezone.now().date()
            motivo = request.POST.get('motivo', '')
            obs_manejo = request.POST.get('obs_manejo', '')
            lote = Lote.objects.get(id=lote_id, user=request.user)
            MovimentacaoLote.objects.create(
                lote=lote,
                data=data,
                tipo=tipo,
                quantidade=quantidade,
                peso_medio=peso_medio,
                motivo=motivo,
                obs_manejo=obs_manejo,
                user=request.user
            )
            messages.success(request, f'Movimentação registrada com sucesso!')
            return redirect('historico_movimentacao_lote')
        except Exception as e:
            messages.error(request, f'Erro ao registrar movimentação: {str(e)}')
    lotes = Lote.objects.filter(user=request.user)
    return render(request, 'estoque/movimentacao_lote.html', {'lotes': lotes})

@login_required
def historico_movimentacao_lote(request):
    lotes = Lote.objects.filter(user=request.user)
    historico = MovimentacaoLote.objects.filter(user=request.user).select_related('lote').order_by('-data')
    return render(request, 'estoque/historico_movimentacao_lote.html', {'lotes': lotes, 'historico': historico})

@login_required
def relatorio_movimentacao_lote(request):
    lotes = Lote.objects.filter(user=request.user)
    relatorio = []
    for lote in lotes:
        movs = MovimentacaoLote.objects.filter(lote=lote, user=request.user).order_by('data')
        estoque_inicial = 0
        estoque_atual = 0
        total_peso = 0
        total_animais = 0
        obs_manejo = ''
        for mov in movs:
            if mov.tipo == 'entrada':
                estoque_atual += mov.quantidade
                total_peso += mov.peso_medio * mov.quantidade
                total_animais += mov.quantidade
            elif mov.tipo in ['saida', 'baixa']:
                estoque_atual -= mov.quantidade
            if mov.obs_manejo:
                obs_manejo = mov.obs_manejo
        media_peso = (total_peso / total_animais) if total_animais > 0 else 0
        relatorio.append({
            'lote': lote,
            'estoque_inicial': 0,  # Pode ser ajustado para considerar um período
            'estoque_atual': estoque_atual,
            'media_peso': round(media_peso, 2),
            'movimentacoes': movs,
            'obs_manejo': obs_manejo,
        })
    return render(request, 'estoque/relatorio_movimentacao_lote.html', {'relatorio': relatorio})

@login_required
def exportar_excel_relatorio_movimentacao_lote(request):
    lotes = Lote.objects.filter(user=request.user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Lotes"
    
    # Cabeçalho seguindo o modelo da planilha
    headers1 = ["VOLUME / CATEG.", "", "", "MOVIMENTAÇÃO", "", "", "", "", "BAIXAS / MOTIVOS", "", "", "ESTOQ. ATUAL", "", "OBS. MANEJO"]
    headers2 = ["LOTES", "ESTOQ", "MED. @", "DATA", "ENTR", "DATA", "ENTR", "SAÍDA", "DATA", "UND", "MOTIV", "DATA", "ESTOQ", "VACINA / VERMIFUGO"]
    ws.append(headers1)
    ws.append(headers2)
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill1 = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Azul escuro
    header_fill2 = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Azul claro
    
    # Aplicar estilo ao primeiro cabeçalho (azul escuro)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill1
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Aplicar estilo ao segundo cabeçalho (azul claro)
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill2
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados dos lotes
    total_estoque_inicial = 0
    total_estoque_atual = 0
    
    for lote in lotes:
        movs = MovimentacaoLote.objects.filter(lote=lote, user=request.user).order_by('data')
        estoque_inicial = 0
        estoque_atual = 0
        total_peso = 0
        total_animais = 0
        obs_manejo = ''
        entradas = []
        saidas = []
        baixas = []
        
        for mov in movs:
            if mov.tipo == 'entrada':
                estoque_atual += mov.quantidade
                total_peso += mov.peso_medio * mov.quantidade
                total_animais += mov.quantidade
                entradas.append(mov)
            elif mov.tipo == 'saida':
                estoque_atual -= mov.quantidade
                saidas.append(mov)
            elif mov.tipo == 'baixa':
                estoque_atual -= mov.quantidade
                baixas.append(mov)
            if mov.obs_manejo:
                obs_manejo = mov.obs_manejo
        
        media_peso = (total_peso / total_animais) if total_animais > 0 else 0
        estoque_inicial = estoque_atual  # Simplificado para o exemplo
        
        # Determinar categoria baseada no peso médio
        if media_peso <= 7:
            categoria = "05 A 07"
        elif media_peso <= 8.5:
            categoria = "8,5"
        elif media_peso <= 9:
            categoria = "08,5 A 09"
        elif media_peso <= 10:
            categoria = "> 10"
        elif media_peso <= 12:
            categoria = "> 12"
        else:
            categoria = "> 12"
        
        # Dados de movimentação
        entrada1_data = entradas[0].data.strftime('%d.%m.%y') if len(entradas) > 0 else ''
        entrada1_qtd = entradas[0].quantidade if len(entradas) > 0 else ''
        entrada2_data = entradas[1].data.strftime('%d.%m.%y') if len(entradas) > 1 else ''
        entrada2_qtd = entradas[1].quantidade if len(entradas) > 1 else ''
        saida_qtd = saidas[0].quantidade if len(saidas) > 0 else ''
        
        # Dados de baixas
        baixa_data = baixas[0].data.strftime('%d.%m.%y') if len(baixas) > 0 else ''
        baixa_qtd = baixas[0].quantidade if len(baixas) > 0 else ''
        baixa_motivo = baixas[0].motivo if len(baixas) > 0 else ''
        
        # Data atual do estoque
        data_atual = timezone.now().strftime('%d.%m.%y')
        
        # Observações de manejo (exemplo)
        obs_manejo = "2ª QUINZ. 07.24 / PROX. 11.24"
        
        row = [
            lote.nome,
            estoque_inicial,
            categoria,
            entrada1_data,
            entrada1_qtd,
            entrada2_data,
            entrada2_qtd,
            saida_qtd,
            baixa_data,
            baixa_qtd,
            baixa_motivo,
            data_atual,
            estoque_atual,
            obs_manejo
        ]
        ws.append(row)
        
        total_estoque_inicial += estoque_inicial
        total_estoque_atual += estoque_atual
    
    # Linha de TOTAL
    total_row = [
        "TOTAL",
        total_estoque_inicial,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        total_estoque_atual,
        ""
    ]
    ws.append(total_row)
    
    # Adicionar seção de ESTOQUE
    ws.append([])  # Linha em branco
    ws.append(["ESTOQUE"])
    ws.append(["ITEM", "DATA DE ENTRADA", "QUANTIDADE", "VALOR UNITÁRIO", "VALOR TOTAL", "QUANTIDADE USADA (MÊS/ANO)"])
    
    # Estilo para cabeçalho do estoque
    estoque_header_font = Font(bold=True, color="FFFFFF")
    estoque_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Verde
    
    for cell in ws[ws.max_row]:
        cell.font = estoque_header_font
        cell.fill = estoque_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados do estoque
    produtos = Produto.objects.filter(user=request.user, quantidade__gt=0)
    for produto in produtos:
        # Buscar movimentações do produto
        movimentacoes = Movimentacao_estoque.objects.filter(produto=produto, user=request.user).order_by('data')
        
        # Calcular quantidade usada no mês atual
        mes_atual = timezone.now().month
        ano_atual = timezone.now().year
        quantidade_usada_mes = movimentacoes.filter(
            tipo='saida',
            data__month=mes_atual,
            data__year=ano_atual
        ).aggregate(total=Sum('quantidade'))['total'] or 0
        
        estoque_row = [
            produto.nome,
            produto.data_de_cadastro.strftime('%d/%m/%Y') if produto.data_de_cadastro else '',
            produto.quantidade,
            f"R$ {produto.valor_unitario:.2f}",
            f"R$ {produto.quantidade * produto.valor_unitario:.2f}",
            f"{quantidade_usada_mes} ({mes_atual}/{ano_atual})"
        ]
        ws.append(estoque_row)
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    # Salvar e retornar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_lotes_estoque.xlsx"'
    return response

@login_required
def exportar_pdf_relatorio_movimentacao_lote(request):
    lotes = Lote.objects.filter(user=request.user)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=20, alignment=1)
    elements.append(Paragraph("Relatório de Movimentação de Lote e Estoque", title_style))
    elements.append(Spacer(1, 12))
    
    total_geral = 0
    total_estoque_inicial = 0
    total_estoque_atual = 0
    
    # Cabeçalho da tabela de lotes
    lote_headers = [
        "VOLUME / CATEG.", "", "", "MOVIMENTAÇÃO", "", "", "", "", "BAIXAS / MOTIVOS", "", "", "ESTOQ. ATUAL", "", "OBS. MANEJO"
    ]
    lote_subheaders = [
        "LOTES", "ESTOQ", "MED. @", "DATA", "ENTR", "DATA", "ENTR", "SAÍDA", "DATA", "UND", "MOTIV", "DATA", "ESTOQ", "VACINA / VERMIFUGO"
    ]
    
    # Dados dos lotes
    lote_data = [lote_headers, lote_subheaders]
    
    for lote in lotes:
        movs = MovimentacaoLote.objects.filter(lote=lote, user=request.user).order_by('data')
        estoque_inicial = 0
        estoque_atual = 0
        total_peso = 0
        total_animais = 0
        obs_manejo = ''
        entradas = []
        saidas = []
        baixas = []
        
        for mov in movs:
            if mov.tipo == 'entrada':
                estoque_atual += mov.quantidade
                total_peso += mov.peso_medio * mov.quantidade
                total_animais += mov.quantidade
                entradas.append(mov)
            elif mov.tipo == 'saida':
                estoque_atual -= mov.quantidade
                saidas.append(mov)
            elif mov.tipo == 'baixa':
                estoque_atual -= mov.quantidade
                baixas.append(mov)
            if mov.obs_manejo:
                obs_manejo = mov.obs_manejo
        
        media_peso = (total_peso / total_animais) if total_animais > 0 else 0
        estoque_inicial = estoque_atual  # Simplificado para o exemplo
        
        # Determinar categoria baseada no peso médio
        if media_peso <= 7:
            categoria = "05 A 07"
        elif media_peso <= 8.5:
            categoria = "8,5"
        elif media_peso <= 9:
            categoria = "08,5 A 09"
        elif media_peso <= 10:
            categoria = "> 10"
        elif media_peso <= 12:
            categoria = "> 12"
        else:
            categoria = "> 12"
        
        # Dados de movimentação
        entrada1_data = entradas[0].data.strftime('%d.%m.%y') if len(entradas) > 0 else ''
        entrada1_qtd = entradas[0].quantidade if len(entradas) > 0 else ''
        entrada2_data = entradas[1].data.strftime('%d.%m.%y') if len(entradas) > 1 else ''
        entrada2_qtd = entradas[1].quantidade if len(entradas) > 1 else ''
        saida_qtd = saidas[0].quantidade if len(saidas) > 0 else ''
        
        # Dados de baixas
        baixa_data = baixas[0].data.strftime('%d.%m.%y') if len(baixas) > 0 else ''
        baixa_qtd = baixas[0].quantidade if len(baixas) > 0 else ''
        baixa_motivo = baixas[0].motivo if len(baixas) > 0 else ''
        
        # Data atual do estoque
        data_atual = timezone.now().strftime('%d.%m.%y')
        
        # Observações de manejo (exemplo)
        obs_manejo = "2ª QUINZ. 07.24 / PROX. 11.24"
        
        row = [
            lote.nome,
            estoque_inicial,
            categoria,
            entrada1_data,
            entrada1_qtd,
            entrada2_data,
            entrada2_qtd,
            saida_qtd,
            baixa_data,
            baixa_qtd,
            baixa_motivo,
            data_atual,
            estoque_atual,
            obs_manejo
        ]
        lote_data.append(row)
        
        total_estoque_inicial += estoque_inicial
        total_estoque_atual += estoque_atual
        total_geral += estoque_atual
    
    # Linha de TOTAL
    total_row = [
        "TOTAL",
        total_estoque_inicial,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        total_estoque_atual,
        ""
    ]
    lote_data.append(total_row)
    
    # Montar tabela dos lotes
    table = Table(lote_data, repeatRows=2)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),  # Azul escuro
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#4F81BD')),  # Azul claro
        ('TEXTCOLOR', (0, 0), (-1, 1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 1), 8),
        ('BACKGROUND', (0, 2), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 2), (-1, -1), 7),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Seção de ESTOQUE
    elements.append(Paragraph("ESTOQUE", title_style))
    elements.append(Spacer(1, 12))
    
    # Cabeçalho do estoque
    estoque_headers = ["ITEM", "DATA DE ENTRADA", "QUANTIDADE", "VALOR UNITÁRIO", "VALOR TOTAL", "QUANTIDADE USADA (MÊS/ANO)"]
    estoque_data = [estoque_headers]
    
    # Dados do estoque
    produtos = Produto.objects.filter(user=request.user, quantidade__gt=0)
    for produto in produtos:
        # Buscar movimentações do produto
        movimentacoes = Movimentacao_estoque.objects.filter(produto=produto, user=request.user).order_by('data')
        
        # Calcular quantidade usada no mês atual
        mes_atual = timezone.now().month
        ano_atual = timezone.now().year
        quantidade_usada_mes = movimentacoes.filter(
            tipo='saida',
            data__month=mes_atual,
            data__year=ano_atual
        ).aggregate(total=Sum('quantidade'))['total'] or 0
        
        estoque_row = [
            produto.nome,
            produto.data_de_cadastro.strftime('%d/%m/%Y') if produto.data_de_cadastro else '',
            produto.quantidade,
            f"R$ {produto.valor_unitario:.2f}",
            f"R$ {produto.quantidade * produto.valor_unitario:.2f}",
            f"{quantidade_usada_mes} ({mes_atual}/{ano_atual})"
        ]
        estoque_data.append(estoque_row)
    
    # Montar tabela do estoque
    estoque_table = Table(estoque_data, repeatRows=1)
    estoque_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),  # Verde
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(estoque_table)
    
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_lotes_estoque.pdf"'
    return response
