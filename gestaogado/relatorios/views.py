# pyright: reportAttributeAccessIssue=false
from django.shortcuts import render
from django.db.models import Count, Sum, Avg
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from datetime import datetime, timedelta
from gado.models import Animal, Vacinacao
from fazendascdst.models import Fazenda, Lote
from estoque.models import Produto, Movimentacao_estoque
from .models import ObservacaoRelatorio
from django.contrib.auth.decorators import login_required
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def index(request):
    # Data atual
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)  # Início do mês atual
    inicio_ano = hoje.replace(month=1, day=1)
    
    # Estatísticas gerais
    total_animais = Animal.objects.filter(user=request.user).count()
    total_fazendas = Fazenda.objects.filter(user=request.user).count()
    total_lotes = Lote.objects.filter(user=request.user).count()
    total_produtos = Produto.objects.filter(user=request.user).count()
    
    # Estatísticas por tipo de gado
    gado_por_tipo = Animal.objects.filter(user=request.user).values('tipo').annotate(
        total=Count('id'),
        peso_medio=Avg('peso')
    )
    
    # Animais por fazenda
    animais_por_fazenda = Animal.objects.filter(user=request.user).values('fazenda__nome').annotate(
        total=Count('id'),
        peso_total=Sum('peso')
    ).order_by('-total')[:5]
    
    # Vacinações do mês atual
    vacinacoes_mes = Vacinacao.objects.filter(
        data_de_aplicacao__gte=inicio_mes, user=request.user
    ).count()
    
    # Animais adicionados este mês
    animais_mes = Animal.objects.filter(
        data_entrada__gte=inicio_mes, user=request.user, is_active=True
    ).count()
    
    # Gastos do mês (movimentações de entrada de estoque)
    gastos_mes = Movimentacao_estoque.objects.filter(
        data__gte=inicio_mes,
        tipo='entrada',
        user=request.user
    ).aggregate(
        total=Sum('quantidade', field='quantidade * valor_unitario')
    )['total'] or 0
    
    # Calcular gastos corretamente
    gastos_mes_corrigido = 0
    movimentacoes_entrada = Movimentacao_estoque.objects.filter(
        data__gte=inicio_mes,
        tipo='entrada',
        user=request.user
    ).select_related('produto')
    
    for mov in movimentacoes_entrada:
        gastos_mes_corrigido += mov.quantidade * mov.produto.valor_unitario
    
    # Produtividade (baseada em atividades do mês)
    atividades_mes = 0
    if animais_mes > 0:
        atividades_mes += 1
    if vacinacoes_mes > 0:
        atividades_mes += 1
    if gastos_mes_corrigido > 0:
        atividades_mes += 1
    
    # Calcular porcentagem de produtividade (baseado em 3 atividades possíveis)
    produtividade_porcentagem = (atividades_mes / 3) * 100 if atividades_mes > 0 else 0
    
    # Atividades recentes (criação de lote, vacinação, entrada/saída de produtos e animais)
    atividades = []
    
    # Adicionar lotes criados recentemente
    lotes_recentes = Lote.objects.filter(user=request.user, is_active=True).order_by('-data_de_cadastro')[:3]
    for lote in lotes_recentes:
        # Contar quantos animais entraram no lote
        animais_no_lote = lote.animais.filter(is_active=True, user=request.user).count()
        atividades.append({
            'objeto': lote,
            'tipo': 'criacao_lote',
            'data': lote.data_de_cadastro,
            'nome': f"Criação do lote {lote.nome}",
            'fazenda': lote.fazenda.nome,
            'descricao': f"{animais_no_lote} animais"
        })
    
    # Adicionar vacinações recentes
    vacinacoes_recentes = Vacinacao.objects.filter(user=request.user).order_by('-data_de_aplicacao')[:3]
    for vacinacao in vacinacoes_recentes:
        atividades.append({
            'objeto': vacinacao,
            'tipo': 'vacinacao',
            'data': vacinacao.data_de_aplicacao,
            'nome': f"Vacinação: {vacinacao.vacina.nome}",
            'fazenda': vacinacao.animal.fazenda.nome,
            'descricao': f"Animal: {vacinacao.animal.identificador or 'SN'}"
        })
    
    # Adicionar movimentações de estoque (entrada/saída de produtos)
    movimentacoes_recentes = Movimentacao_estoque.objects.filter(user=request.user).order_by('-data')[:3]
    for movimentacao in movimentacoes_recentes:
        tipo_mov = "Entrada" if movimentacao.tipo == 'entrada' else "Saída"
        atividades.append({
            'objeto': movimentacao,
            'tipo': 'movimentacao_estoque',
            'data': movimentacao.data,
            'nome': f"{tipo_mov} de produto",
            'fazenda': movimentacao.produto.fazenda.nome,
            'descricao': f"{movimentacao.produto.nome} - {movimentacao.quantidade} unidades"
        })
    
    # Adicionar animais com entrada recente
    animais_recentes = Animal.objects.filter(user=request.user, is_active=True).order_by('-data_entrada')[:3]
    for animal in animais_recentes:
        atividades.append({
            'objeto': animal,
            'tipo': 'entrada_animal',
            'data': animal.data_entrada,
            'nome': f"Entrada de animal",
            'fazenda': animal.fazenda.nome,
            'descricao': f"Animal: {animal.identificador or 'SN'} - {animal.tipo}"
        })
    
    # Ordenar por data (mais recentes primeiro)
    atividades.sort(key=lambda x: x['data'], reverse=True)
    
    # Pegar apenas os primeiros 8
    atividades_final = atividades[:8]
    
    context = {
        'total_animais': total_animais,
        'total_fazendas': total_fazendas,
        'total_lotes': total_lotes,
        'total_produtos': total_produtos,
        'atividades_recentes': atividades_final,
        'vacinacoes_mes': vacinacoes_mes,
        'animais_mes': animais_mes,
        'gastos_mes': gastos_mes_corrigido,
        'produtividade_porcentagem': produtividade_porcentagem,
        'mes_atual': f"Mês de {inicio_mes.strftime('%B/%Y')}",
    }
    return render(request, 'relatorios/index.html', context)

def relatorio_mensal(request):
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    
    # Dados do mês atual
    animais_mes = Animal.objects.filter(data_entrada__gte=inicio_mes, user=request.user)
    vacinacoes_mes = Vacinacao.objects.filter(data_de_aplicacao__gte=inicio_mes, user=request.user)
    
    # Movimentações de estoque do mês
    movimentacoes_mes = Movimentacao_estoque.objects.filter(data__gte=inicio_mes, user=request.user)
    
    # Calcular investimentos do mês (movimentações de entrada)
    investimentos_mes = 0
    movimentacoes_entrada = Movimentacao_estoque.objects.filter(
        data__gte=inicio_mes,
        tipo='entrada',
        user=request.user
    ).select_related('produto')
    
    for mov in movimentacoes_entrada:
        if mov.produto and mov.produto.valor_unitario:
            investimentos_mes += mov.quantidade * mov.produto.valor_unitario
    
    context = {
        'mes_atual': hoje.strftime('%B %Y'),
        'animais_mes': animais_mes,
        'vacinacoes_mes': vacinacoes_mes,
        'movimentacoes_mes': movimentacoes_mes,
        'total_animais_mes': animais_mes.count(),
        'total_vacinacoes_mes': vacinacoes_mes.count(),
        'investimentos_mes': investimentos_mes,
    }
    
    return render(request, 'relatorios/mensal.html', context)

def relatorio_anual(request):
    hoje = timezone.now().date()
    inicio_ano = hoje.replace(month=1, day=1)
    
    # Dados do ano atual
    animais_ano = Animal.objects.filter(data_entrada__gte=inicio_ano, user=request.user)
    vacinacoes_ano = Vacinacao.objects.filter(data_de_aplicacao__gte=inicio_ano, user=request.user)
    
    # Estatísticas anuais por mês
    dados_por_mes = []
    for mes in range(1, 13):
        # Animais do mês
        animais_mes = Animal.objects.filter(
            data_entrada__year=hoje.year,
            data_entrada__month=mes,
            user=request.user
        ).count()
        
        # Vacinações do mês
        vacinacoes_mes = Vacinacao.objects.filter(
            data_de_aplicacao__year=hoje.year,
            data_de_aplicacao__month=mes,
            user=request.user
        ).count()
        
        # Gastos do mês (movimentações de entrada)
        gastos_mes = 0
        movimentacoes_mes = Movimentacao_estoque.objects.filter(
            data__year=hoje.year,
            data__month=mes,
            tipo='entrada',
            user=request.user
        ).select_related('produto')
        
        for mov in movimentacoes_mes:
            if mov.produto and mov.produto.valor_unitario:
                gastos_mes += mov.quantidade * mov.produto.valor_unitario
        
        # Buscar observação do mês
        observacao_obj = ObservacaoRelatorio.get_observacao(request.user, hoje.year, mes)
        observacao_texto = observacao_obj.observacao if observacao_obj else ''
        
        dados_por_mes.append({
            'mes': mes,
            'nome_mes': datetime(hoje.year, mes, 1).strftime('%B'),
            'animais': animais_mes,
            'vacinacoes': vacinacoes_mes,
            'gastos': gastos_mes,
            'observacao': observacao_texto
        })
    
    # Calcular investimento total (movimentações de entrada de estoque)
    investimento_total = 0
    movimentacoes_entrada = Movimentacao_estoque.objects.filter(
        data__gte=inicio_ano,
        tipo='entrada',
        user=request.user
    ).select_related('produto')
    
    for mov in movimentacoes_entrada:
        if mov.produto and mov.produto.valor_unitario:
            investimento_total += mov.quantidade * mov.produto.valor_unitario
    
    # Total de produtos em estoque
    total_produtos_estoque = Produto.objects.filter(user=request.user).count()
    
    # Total de fazendas ativas
    total_fazendas_ativas = Fazenda.objects.filter(user=request.user, is_active=True).count()
    
    context = {
        'ano_atual': hoje.year,
        'animais_ano': animais_ano,
        'vacinacoes_ano': vacinacoes_ano,
        'total_animais_ano': animais_ano.count(),
        'total_vacinacoes_ano': vacinacoes_ano.count(),
        'investimento_total': investimento_total,
        'total_produtos_estoque': total_produtos_estoque,
        'total_fazendas_ativas': total_fazendas_ativas,
        'dados_por_mes': dados_por_mes,
    }
    
    return render(request, 'relatorios/anual.html', context)

# Funções de exportação
def exportar_pdf_relatorio_mensal(request):
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    
    # Dados do mês atual
    animais_mes = Animal.objects.filter(data_entrada__gte=inicio_mes, user=request.user)
    vacinacoes_mes = Vacinacao.objects.filter(data_de_aplicacao__gte=inicio_mes, user=request.user)
    movimentacoes_mes = Movimentacao_estoque.objects.filter(data__gte=inicio_mes, user=request.user)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, alignment=1)
    section_title = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=13, spaceAfter=12, alignment=0)
    
    # Título principal
    elements.append(Paragraph(f"Relatório Mensal - {hoje.strftime('%B %Y')}", title_style))
    elements.append(Spacer(1, 20))
    
    # 1. Animais Adicionados
    elements.append(Paragraph("Animais Adicionados Este Mês", section_title))
    animais_data = [['Identificador', 'Data de Entrada', 'Peso (kg)', 'Fazenda']]
    for animal in animais_mes:
        animais_data.append([
            str(animal.identificador or 'SN'),
            animal.data_entrada.strftime('%d/%m/%Y'),
            f"{animal.peso:.1f}kg",
            animal.fazenda.nome
        ])
    if len(animais_data) == 1:
        animais_data.append(['-', '-', '-', '-'])
    animais_table = Table(animais_data, repeatRows=1)
    animais_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(animais_table)
    elements.append(PageBreak())
    
    # 2. Vacinações Realizadas
    elements.append(Paragraph("Vacinações Realizadas", section_title))
    vacinas_data = [['Animal', 'Vacina', 'Data', 'Status']]
    for vac in vacinacoes_mes:
        vacinas_data.append([
            str(vac.animal.identificador or 'SN'),
            vac.vacina.nome,
            vac.data_de_aplicacao.strftime('%d/%m/%Y'),
            'Realizada'
        ])
    if len(vacinas_data) == 1:
        vacinas_data.append(['-', '-', '-', '-'])
    vacinas_table = Table(vacinas_data, repeatRows=1)
    vacinas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(vacinas_table)
    elements.append(PageBreak())
    
    # 3. Movimentação de Estoque
    elements.append(Paragraph("Movimentação de Estoque", section_title))
    mov_data = [['Produto', 'Tipo de Movimento', 'Quantidade', 'Data', 'Valor Unitário', 'Total']]
    for mov in movimentacoes_mes:
        valor_unitario = float(mov.produto.valor_unitario) if mov.produto and mov.produto.valor_unitario else 0
        total = float(mov.quantidade) * valor_unitario
        mov_data.append([
            mov.produto.nome if mov.produto else 'N/A',
            mov.get_tipo_display() if hasattr(mov, 'get_tipo_display') else str(mov.tipo),
            mov.quantidade,
            mov.data.strftime('%d/%m/%Y') if mov.data else '',
            f"R$ {valor_unitario:.2f}",
            f"R$ {total:.2f}"
        ])
    if len(mov_data) == 1:
        mov_data.append(['-', '-', '-', '-', '-', '-'])
    mov_table = Table(mov_data, repeatRows=1)
    mov_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(mov_table)
    
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_mensal_{hoje.strftime("%Y_%m")}.pdf"'
    return response

def exportar_excel_relatorio_mensal(request):
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    
    # Dados do mês atual
    animais_mes = Animal.objects.filter(data_entrada__gte=inicio_mes, user=request.user)
    vacinacoes_mes = Vacinacao.objects.filter(data_de_aplicacao__gte=inicio_mes, user=request.user)
    movimentacoes_mes = Movimentacao_estoque.objects.filter(data__gte=inicio_mes, user=request.user)
    
    wb = Workbook()
    
    # Aba 1: Animais Adicionados
    ws1 = wb.active
    ws1.title = "Animais Adicionados"
    ws1['A1'] = f"Animais Adicionados - {hoje.strftime('%B %Y')}"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:D1')
    ws1['A3'] = "Identificador"
    ws1['B3'] = "Data de Entrada"
    ws1['C3'] = "Peso (kg)"
    ws1['D3'] = "Fazenda"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws1[3]:
        cell.font = header_font
        cell.fill = header_fill
    row = 4
    for animal in animais_mes:
        ws1[f'A{row}'] = animal.identificador or 'SN'
        ws1[f'B{row}'] = animal.data_entrada.strftime('%d/%m/%Y')
        ws1[f'C{row}'] = float(animal.peso)
        ws1[f'D{row}'] = animal.fazenda.nome
        row += 1
    for column in ws1.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Aba 2: Vacinações Realizadas
    ws2 = wb.create_sheet("Vacinações Realizadas")
    ws2['A1'] = f"Vacinações Realizadas - {hoje.strftime('%B %Y')}"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:D1')
    ws2['A3'] = "Animal"
    ws2['B3'] = "Vacina"
    ws2['C3'] = "Data"
    ws2['D3'] = "Status"
    for cell in ws2[3]:
        cell.font = header_font
        cell.fill = header_fill
    row = 4
    for vac in vacinacoes_mes:
        ws2[f'A{row}'] = vac.animal.identificador or 'SN'
        ws2[f'B{row}'] = vac.vacina.nome
        ws2[f'C{row}'] = vac.data_de_aplicacao.strftime('%d/%m/%Y')
        ws2[f'D{row}'] = 'Realizada'
        row += 1
    for column in ws2.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Aba 3: Movimentação de Estoque
    ws3 = wb.create_sheet("Movimentação de Estoque")
    ws3['A1'] = f"Movimentação de Estoque - {hoje.strftime('%B %Y')}"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.merge_cells('A1:F1')
    ws3['A3'] = "Produto"
    ws3['B3'] = "Tipo de Movimento"
    ws3['C3'] = "Quantidade"
    ws3['D3'] = "Data"
    ws3['E3'] = "Valor Unitário"
    ws3['F3'] = "Total"
    for cell in ws3[3]:
        cell.font = header_font
        cell.fill = header_fill
    row = 4
    for mov in movimentacoes_mes:
        ws3[f'A{row}'] = mov.produto.nome if mov.produto else 'N/A'
        ws3[f'B{row}'] = mov.get_tipo_display() if hasattr(mov, 'get_tipo_display') else str(mov.tipo)
        ws3[f'C{row}'] = mov.quantidade
        ws3[f'D{row}'] = mov.data.strftime('%d/%m/%Y') if mov.data else ''
        ws3[f'E{row}'] = float(mov.produto.valor_unitario) if mov.produto and mov.produto.valor_unitario else 0
        ws3[f'F{row}'] = float(mov.quantidade) * float(mov.produto.valor_unitario) if mov.produto and mov.produto.valor_unitario else float(mov.quantidade)
        row += 1
    for column in ws3.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws3.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_mensal_{hoje.strftime("%Y_%m")}.xlsx"'
    return response

def exportar_pdf_relatorio_anual(request):
    hoje = timezone.now().date()
    inicio_ano = hoje.replace(month=1, day=1)
    
    # Criar PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1
    )
    
    # Título
    elements.append(Paragraph(f"Relatório Anual - {hoje.year}", title_style))
    elements.append(Spacer(1, 20))
    
    # Seção de Controle de Lotes
    elements.append(Paragraph("Controle de Lotes", styles['Heading2']))
    lotes_data = [['Lote', 'Data Entrada', 'Data Saída', 'Baixas/Motivos', 'Estoque Atual', 'Lote de Saída']]
    
    # Obter todos os lotes com animais
    lotes = Lote.objects.filter(
        animais__user=request.user,
        animais__is_active=True
    ).distinct()
    
    total_geral_animais = 0
    
    for lote in lotes:
        # Animais no lote
        animais_no_lote = lote.animais.filter(user=request.user, is_active=True)
        quantidade_atual = animais_no_lote.count()
        total_geral_animais += quantidade_atual
        
        # Animais que saíram do lote no ano
        animais_saida = lote.animais.filter(
            user=request.user,
            data_de_saida__gte=inicio_ano,
            data_de_saida__lte=hoje
        )
        
        # Data de entrada (primeiro animal)
        primeiro_animal = animais_no_lote.order_by('data_entrada').first()
        data_entrada = primeiro_animal.data_entrada.strftime('%d/%m/%Y') if primeiro_animal else 'N/A'
        
        # Data de saída (último animal que saiu)
        ultimo_saida = animais_saida.order_by('-data_de_saida').first()
        data_saida = ultimo_saida.data_de_saida.strftime('%d/%m/%Y') if ultimo_saida else 'N/A'
        
        # Observações de baixa (primeira baixa do ano)
        observacoes = ''
        if animais_saida.exists():
            primeira_baixa = animais_saida.filter(observacoes_baixa__isnull=False).exclude(observacoes_baixa='').first()
            if primeira_baixa:
                observacoes = primeira_baixa.observacoes_baixa[:50] + '...' if len(primeira_baixa.observacoes_baixa) > 50 else primeira_baixa.observacoes_baixa
        
        # Lote de saída (para onde foi movido)
        lote_saida = 'N/A'
        if ultimo_saida and ultimo_saida.lote and ultimo_saida.lote != lote:
            lote_saida = ultimo_saida.lote.nome
        
        lotes_data.append([
            lote.nome,
            data_entrada,
            data_saida,
            observacoes,
            str(quantidade_atual),
            lote_saida
        ])
    
    # Adicionar linha de total
    lotes_data.append(['TOTAL', '', '', '', str(total_geral_animais), ''])
    
    lotes_table = Table(lotes_data)
    lotes_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(lotes_table)
    elements.append(Spacer(1, 20))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_anual_{hoje.year}.pdf"'
    return response

def exportar_excel_relatorio_anual(request):
    hoje = timezone.now().date()
    inicio_ano = hoje.replace(month=1, day=1)
    
    # Criar workbook
    wb = Workbook()
    
    # Planilha de controle de lotes
    ws1 = wb.active
    ws1.title = "Controle de Lotes"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Título
    ws1['A1'] = f"Relatório Anual - {hoje.year}"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:F1')
    
    # Cabeçalho da tabela
    ws1['A3'] = "Lote"
    ws1['B3'] = "Data Entrada"
    ws1['C3'] = "Data Saída"
    ws1['D3'] = "Baixas/Motivos"
    ws1['E3'] = "Estoque Atual"
    ws1['F3'] = "Lote de Saída"
    
    # Aplicar estilo ao cabeçalho
    for cell in ws1[3]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Obter todos os lotes com animais
    lotes = Lote.objects.filter(
        animais__user=request.user,
        animais__is_active=True
    ).distinct()
    
    # Dados dos lotes
    row = 4
    total_geral_animais = 0
    for lote in lotes:
        # Animais no lote
        animais_no_lote = lote.animais.filter(user=request.user, is_active=True)
        quantidade_atual = animais_no_lote.count()
        total_geral_animais += quantidade_atual
        
        # Animais que saíram do lote no ano
        animais_saida = lote.animais.filter(
            user=request.user,
            data_de_saida__gte=inicio_ano,
            data_de_saida__lte=hoje
        )
        
        # Data de entrada (primeiro animal)
        primeiro_animal = animais_no_lote.order_by('data_entrada').first()
        data_entrada = primeiro_animal.data_entrada.strftime('%d/%m/%Y') if primeiro_animal else 'N/A'
        
        # Data de saída (último animal que saiu)
        ultimo_saida = animais_saida.order_by('-data_de_saida').first()
        data_saida = ultimo_saida.data_de_saida.strftime('%d/%m/%Y') if ultimo_saida else 'N/A'
        
        # Observações de baixa (primeira baixa do ano)
        observacoes = ''
        if animais_saida.exists():
            primeira_baixa = animais_saida.filter(observacoes_baixa__isnull=False).exclude(observacoes_baixa='').first()
            if primeira_baixa:
                observacoes = primeira_baixa.observacoes_baixa
        
        # Lote de saída (para onde foi movido)
        lote_saida = 'N/A'
        if ultimo_saida and ultimo_saida.lote and ultimo_saida.lote != lote:
            lote_saida = ultimo_saida.lote.nome
        
        ws1[f'A{row}'] = lote.nome
        ws1[f'B{row}'] = data_entrada
        ws1[f'C{row}'] = data_saida
        ws1[f'D{row}'] = observacoes
        ws1[f'E{row}'] = quantidade_atual
        ws1[f'F{row}'] = lote_saida
        row += 1
    
    # Adicionar linha de total
    ws1[f'A{row}'] = 'TOTAL'
    ws1[f'E{row}'] = total_geral_animais
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'E{row}'].font = Font(bold=True)
    
    # Ajustar largura das colunas
    for column in ws1.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_anual_{hoje.year}.xlsx"'
    return response

@login_required
def dashboard(request):
    # Data atual
    hoje = timezone.now().date()
    inicio_semana = hoje - timedelta(days=hoje.weekday())  # Início da semana (segunda-feira)
    
    # Totais
    total_animais = Animal.objects.filter(user=request.user, is_active=True).count()
    total_fazendas = Fazenda.objects.filter(user=request.user, is_active=True).count()
    total_lotes = Lote.objects.filter(user=request.user, is_active=True).count()
    total_produtos = Produto.objects.filter(user=request.user).count()

    # Vacinações da semana atual
    vacinacoes_semana = Vacinacao.objects.filter(
        data_de_aplicacao__gte=inicio_semana, user=request.user
    ).count()

    # Atividades recentes (últimos 5 animais, produtos, lotes)
    atividades = []
    
    # Adicionar animais com data de entrada
    animais_recentes = Animal.objects.filter(user=request.user, is_active=True).order_by('-data_entrada')[:5]
    for animal in animais_recentes:
        atividades.append({
            'objeto': animal,
            'tipo': 'animal',
            'data': animal.data_entrada,
            'nome': str(animal.identificador or 'SN'),
            'fazenda': animal.fazenda.nome
        })
    
    # Adicionar lotes com data de cadastro
    lotes_recentes = Lote.objects.filter(user=request.user, is_active=True).order_by('-data_de_cadastro')[:5]
    for lote in lotes_recentes:
        # Contar quantos animais entraram no lote
        animais_no_lote = lote.animais.filter(is_active=True, user=request.user).count()
        atividades.append({
            'objeto': lote,
            'tipo': 'lote',
            'data': lote.data_de_cadastro,
            'nome': lote.nome,
            'fazenda': lote.fazenda.nome,
            'descricao': f"{animais_no_lote} animais no lote"
        })
    
    # Adicionar produtos (sem data específica, usar ID como referência)
    produtos_recentes = Produto.objects.filter(user=request.user).order_by('-id')[:5]
    for produto in produtos_recentes:
        atividades.append({
            'objeto': produto,
            'tipo': 'produto',
            'data': None,  # Produto não tem data específica
            'nome': produto.nome,
            'fazenda': produto.fazenda.nome
        })
    
    # Ordenar por data (produtos ficam no final por não terem data)
    atividades_com_data = [a for a in atividades if a['data']]
    atividades_sem_data = [a for a in atividades if not a['data']]
    
    atividades_com_data.sort(key=lambda x: x['data'], reverse=True)
    atividades_final = atividades_com_data + atividades_sem_data
    
    # Pegar apenas os primeiros 5
    atividades_final = atividades_final[:5]
    
    context = {
        'total_animais': total_animais,
        'total_fazendas': total_fazendas,
        'total_lotes': total_lotes,
        'total_produtos': total_produtos,
        'atividades_recentes': atividades_final,
        'vacinacoes_semana': vacinacoes_semana,
        'semana_atual': f"Semana de {inicio_semana.strftime('%d/%m/%Y')} a {hoje.strftime('%d/%m/%Y')}",
    }
    return render(request, 'relatorios/index.html', context)

# Views para gerenciar observações dos relatórios
@login_required
def editar_observacao(request, ano, mes):
    """Edita ou cria uma observação para um mês específico"""
    if request.method == 'POST':
        observacao_texto = request.POST.get('observacao', '').strip()
        
        # Buscar ou criar observação
        observacao, created = ObservacaoRelatorio.objects.get_or_create(
            user=request.user,
            ano=ano,
            mes=mes,
            defaults={'observacao': observacao_texto}
        )
        
        if not created:
            observacao.observacao = observacao_texto
            observacao.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Observação salva com sucesso!',
            'observacao': observacao_texto
        })
    
    # GET request - retorna formulário
    observacao_obj = ObservacaoRelatorio.get_observacao(request.user, ano, mes)
    observacao_texto = observacao_obj.observacao if observacao_obj else ''
    
    context = {
        'ano': ano,
        'mes': mes,
        'nome_mes': datetime(ano, mes, 1).strftime('%B'),
        'observacao': observacao_texto
    }
    return render(request, 'relatorios/editar_observacao.html', context)

@login_required
def obter_observacao(request, ano, mes):
    """Retorna a observação para um mês específico via AJAX"""
    observacao_obj = ObservacaoRelatorio.get_observacao(request.user, ano, mes)
    observacao_texto = observacao_obj.observacao if observacao_obj else ''
    
    return JsonResponse({
        'observacao': observacao_texto
    })

# Views para gerenciar observações detalhadas
@login_required
def adicionar_observacao_detalhada(request):
    """Adiciona uma nova observação detalhada"""
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        lote = request.POST.get('lote', '').strip()
        data_referencia = request.POST.get('data_referencia')
        observacao = request.POST.get('observacao', '').strip()
        
        # Se não há observação, não salva o registro
        if not observacao:
            return JsonResponse({'success': False, 'error': 'Observação é obrigatória'})
        
        # Se há observação mas não há tipo, define como geral
        if not tipo:
            tipo = 'GERAL'
        
        hoje = timezone.now().date()
        
        # Converter data se fornecida
        data_ref = None
        if data_referencia:
            try:
                data_ref = datetime.strptime(data_referencia, '%Y-%m-%d').date()
            except:
                pass
        
        ObservacaoRelatorio.objects.create(
            user=request.user,
            tipo=tipo,
            lote=lote if lote else None,
            data_referencia=data_ref,
            observacao=observacao,
            mes=hoje.month,
            ano=hoje.year
        )
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'Dados inválidos'})

@login_required
def listar_observacoes_detalhadas(request):
    """Lista as observações detalhadas do mês atual"""
    hoje = timezone.now().date()
    observacoes = ObservacaoRelatorio.objects.filter(
        user=request.user,
        mes=hoje.month,
        ano=hoje.year
    ).order_by('-data_criacao')
    
    data = []
    for obs in observacoes:
        data.append({
            'id': obs.id,
            'tipo': obs.get_tipo_display(),
            'lote': obs.lote or 'N/A',
            'data_referencia': obs.data_referencia.strftime('%d/%m/%Y') if obs.data_referencia else 'N/A',
            'observacao': obs.observacao,
            'data_criacao': obs.data_criacao.strftime('%d/%m/%Y %H:%M')
        })
    
    return JsonResponse({'observacoes': data})

@login_required
def excluir_observacao_detalhada(request, obs_id):
    """Exclui uma observação detalhada"""
    try:
        obs = ObservacaoRelatorio.objects.get(id=obs_id, user=request.user)
        obs.delete()
        return JsonResponse({'success': True})
    except ObservacaoRelatorio.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Observação não encontrada'})